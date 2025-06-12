# app_pages/p13_revenue_reports.py
import streamlit as st
import pandas as pd
from sqlalchemy.orm import Session, joinedload # <-- IMPORT joinedload
from sqlalchemy import func, and_
from decimal import Decimal
import datetime
import sys
import os
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- Boilerplate: Add project root to path ---
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from models import User, Invoice, PurchaseOrder, Transaction, TransactionType, ExpenseCategory, InvoiceStatus

# ==============================================================================
# --- NEW: Excel Report Generation Logic ---
# ==============================================================================
def generate_excel_report(db: Session, user: User, start_date: datetime.date, end_date: datetime.date) -> io.BytesIO:
    """
    Generates an Excel file mirroring the structure of the provided spreadsheet.
    """
    # 1. Fetch all necessary data in the date range
    # --- FIX: Use joinedload correctly ---
    transactions = db.query(Transaction).options(
        joinedload(Transaction.category_ref)
    ).filter(
        Transaction.user_id == user.id,
        Transaction.date.between(start_date, end_date)
    ).order_by(Transaction.date.asc()).all()

    # 2. Create an in-memory Excel workbook
    wb = Workbook()
    wb.remove(wb.active) # Remove the default sheet

    # 3. Pre-process data and create a mapping for expense columns
    df = pd.DataFrame([{
        'date': t.date,
        'description': t.description,
        'type': t.transaction_type,
        'amount': t.amount,
        'category': t.category_ref.name if t.category_ref else 'Misc'
    } for t in transactions])
    
    if not df.empty:
        df['month_name'] = pd.to_datetime(df['date']).dt.strftime('%B')
    
    expense_categories = [cat.name for cat in db.query(ExpenseCategory).filter(ExpenseCategory.user_id == user.id).order_by(ExpenseCategory.name).all()]
    category_to_column = {cat: i + 4 for i, cat in enumerate(expense_categories)} # D is 4, E is 5 etc.

    # --- Generate Monthly Sheets ---
    months = pd.date_range(start_date, end_date, freq='MS').strftime('%B').unique()
    for month_name in months:
        ws = wb.create_sheet(title=month_name)
        
        headers = ["Date", "Who", "Ref", "Amount"] + expense_categories + ["Total"]
        ws.append(headers)
        
        month_df = df[(df['month_name'] == month_name) & (df['type'] == TransactionType.EXPENSE)]
        
        row_num = 2
        for _, row in month_df.iterrows():
            excel_row = [None] * len(headers)
            excel_row[0] = row['date']
            excel_row[1] = row['description']
            excel_row[3] = row['amount']
            
            col_idx = category_to_column.get(row['category'])
            if col_idx:
                excel_row[col_idx + 1] = row['amount']
            
            excel_row[-1] = f"=SUM(E{row_num}:{get_column_letter(len(headers)-1)}{row_num})"
            ws.append(excel_row)
            row_num += 1

        total_row_idx = 33
        ws.cell(row=total_row_idx, column=1, value="Total")
        for i in range(3, len(headers)):
            col_letter = get_column_letter(i + 1)
            ws.cell(row=total_row_idx, column=i + 1, value=f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})")

    # --- Generate Summary Sheet ---
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["Summary for year"])
    summary_headers = ["", "Amount"] + expense_categories + ["Total", "Sales"]
    ws_summary.append(summary_headers)

    row_num = 4
    for month_name in months:
        month_row = [month_name]
        # These formulas pull the totals from each monthly sheet
        for i in range(len(summary_headers) - 2): # Loop through Amount, categories, Total
             col_letter = get_column_letter(i + 2) # B, C, D...
             month_row.append(f"='{month_name}'!{col_letter}33")
        
        # Add formula for sales for that month
        sales_formula = f"=SUMIFS('Bank Transactions'!E:E, 'Bank Transactions'!A:A, \">=\"&DATE({start_date.year},{pd.to_datetime(month_name, format='%B').month},1), 'Bank Transactions'!A:A, \"<=\"&EOMONTH(DATE({start_date.year},{pd.to_datetime(month_name, format='%B').month},1),0))"
        month_row.append(sales_formula)
        ws_summary.append(month_row)
        row_num += 1
    
    # --- Generate Bank Transactions Sheet ---
    ws_bank = wb.create_sheet(title="Bank Transactions", index=1)
    ws_bank.append(["Date", "Who", "Ref", "Out", "In", "Balance"])

    balance = Decimal('0.0')
    for _, row in df.iterrows():
        is_income = row['type'] in [TransactionType.SALE, TransactionType.CAPITAL_INJECTION]
        amount = row['amount']
        
        out_val = float(amount) if not is_income else None
        in_val = float(amount) if is_income else None
        
        if is_income:
            balance += amount
        else:
            balance -= amount
            
        ws_bank.append([row['date'], row['description'], "", out_val, in_val, float(balance)])

    # 4. Save workbook to a memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==============================================================================
# --- Main Render Function ---
# ==============================================================================
def render(db: Session, user: User, is_mobile: bool):
    st.header("📊 Financial Reports")
    st.write("Generate key financial summaries for tax reporting and business analysis.")

    st.subheader("Select Reporting Period")
    today = datetime.date.today()
    c1, c2 = st.columns(2)
    start_date = c1.date_input("From", today.replace(month=1, day=1))
    end_date = c2.date_input("To", today)

    if start_date > end_date:
        st.error("Start date cannot be after end date."); return

    st.markdown("---")

    if user.country_code == 'IE':
        st.subheader("🇮🇪 VAT 3 Summary (Irish Revenue)")
        st.info("This report summarizes the Value Added Tax (VAT) on your sales and purchases for your selected period, as required for a VAT 3 return.")

        vat_on_sales = db.query(func.sum(Invoice.vat_amount)).filter(
            Invoice.user_id == user.id,
            Invoice.status != InvoiceStatus.VOID,
            Invoice.invoice_date.between(start_date, end_date)
        ).scalar() or Decimal('0.0')

        vat_on_purchases = db.query(func.sum(PurchaseOrder.total_vat)).filter(
            PurchaseOrder.user_id == user.id,
            PurchaseOrder.order_date.between(start_date, end_date)
        ).scalar() or Decimal('0.0')

        net_vat_payable = vat_on_sales - vat_on_purchases

        v1, v2, v3 = st.columns(3)
        v1.metric("T1: VAT on Sales", f"€{vat_on_sales:,.2f}")
        v2.metric("T2: VAT on Purchases", f"€{vat_on_purchases:,.2f}")
        v3.metric("Net VAT Payable / Reclaimable", f"€{net_vat_payable:,.2f}", delta=f"€{-net_vat_payable:,.2f}")
        st.markdown("---")

    st.subheader("Profit & Loss Statement")
    
    transactions = db.query(Transaction).filter(
        Transaction.user_id == user.id,
        Transaction.date.between(start_date, end_date)
    ).all()

    total_sales = sum(t.amount for t in transactions if t.transaction_type == TransactionType.SALE)
    total_expenses = sum(t.amount for t in transactions if t.transaction_type == TransactionType.EXPENSE)
    net_profit = total_sales - total_expenses
    
    p1, p2, p3 = st.columns(3)
    p1.metric("Total Sales Revenue", f"€{total_sales:,.2f}")
    p2.metric("Total Expenses", f"€{-total_expenses:,.2f}")
    p3.metric("Net Profit / Loss", f"€{net_profit:,.2f}", delta=f"{net_profit:,.2f}")

    st.subheader("Expense Breakdown by Category")
    expense_transactions = [t for t in transactions if t.transaction_type == TransactionType.EXPENSE]

    if not expense_transactions:
        st.info("No expenses recorded in this period.")
    else:
        expense_data = {}
        for t in expense_transactions:
            category_name = t.category_ref.name if t.category_ref else "Uncategorized"
            if category_name not in expense_data:
                expense_data[category_name] = Decimal('0.0')
            expense_data[category_name] += t.amount
        
        df_expenses = pd.DataFrame(list(expense_data.items()), columns=['Category', 'Amount'])
        df_expenses = df_expenses.sort_values('Amount', ascending=False)
        
        # --- FIX: Convert 'Amount' to float for charting ---
        df_expenses['Amount'] = df_expenses['Amount'].astype(float)

        c1, c2 = st.columns([1,1])
        with c1:
            st.write("Expense Chart")
            st.bar_chart(df_expenses.set_index('Category'))
        with c2:
            st.write("Expense Data")
            st.dataframe(df_expenses, hide_index=True, use_container_width=True,
                         column_config={"Amount": st.column_config.NumberColumn(format="€%.2f")})

    st.markdown("---")

    st.subheader("Download Full Report")
    if 'excel_data' not in st.session_state:
        st.session_state.excel_data = None

    if st.button("Generate Excel Report"):
        with st.spinner("Generating your Excel report..."):
            st.session_state.excel_data = generate_excel_report(db, user, start_date, end_date)
    
    if st.session_state.excel_data:
        st.download_button(
            label="📥 Download Excel File",
            data=st.session_state.excel_data,
            file_name=f"Financial_Report_{start_date}_to_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )